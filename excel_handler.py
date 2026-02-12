"""
Excel Import/Export Handler for Quality System
Supports importing and exporting records, templates, standards, and reports
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import pandas as pd
from typing import List, Dict, Any
from models import *


class ExcelHandler:
    """Handle Excel import/export operations"""
    
    def __init__(self, session):
        """
        Initialize Excel handler
        
        Args:
            session: SQLAlchemy session
        """
        self.session = session
    
    # ========================================================================
    # EXPORT METHODS
    # ========================================================================
    
    def export_records_to_excel(self, records: List[Record], filepath: str) -> str:
        """
        Export records to Excel file with formatting
        
        Args:
            records: List of Record objects
            filepath: Output file path
            
        Returns:
            Path to created file
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Records"
        
        # Define headers
        headers = [
            'Record Number', 'Title', 'Template', 'Status', 'Priority',
            'Scheduled Date', 'Completed Date', 'Batch Number', 'Product ID',
            'Location', 'Department', 'Created By', 'Assigned To',
            'Compliance Score', 'Failed Items', 'Overall Compliance', 'Notes'
        ]
        
        # Style definitions
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Write data
        for row_idx, record in enumerate(records, 2):
            data = [
                record.record_number,
                record.title,
                record.template.name if record.template else '',
                record.status,
                record.priority or '',
                record.scheduled_date.strftime('%Y-%m-%d %H:%M') if record.scheduled_date else '',
                record.completed_at.strftime('%Y-%m-%d %H:%M') if record.completed_at else '',
                record.batch_number or '',
                record.product_id or '',
                record.location or '',
                record.department or '',
                record.creator.full_name if record.creator else '',
                record.assignee.full_name if record.assignee else '',
                float(record.compliance_score) if record.compliance_score else '',
                record.failed_items_count or 0,
                'Pass' if record.overall_compliance else 'Fail' if record.overall_compliance is not None else '',
                record.notes or ''
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.value = value
                cell.border = border
                
                # Conditional formatting for status
                if col == 4:  # Status column
                    if value == 'approved':
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif value == 'rejected':
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                
                # Conditional formatting for overall compliance
                if col == 16:  # Overall compliance column
                    if value == 'Pass':
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                        cell.font = Font(color='006100', bold=True)
                    elif value == 'Fail':
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                        cell.font = Font(color='9C0006', bold=True)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save file
        wb.save(filepath)
        return filepath
    
    def export_record_details_to_excel(self, record: Record, filepath: str) -> str:
        """
        Export detailed record with all items to Excel
        
        Args:
            record: Record object
            filepath: Output file path
            
        Returns:
            Path to created file
        """
        wb = openpyxl.Workbook()
        
        # Sheet 1: Record Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        summary_data = [
            ['Record Number', record.record_number],
            ['Title', record.title],
            ['Template', record.template.name if record.template else ''],
            ['Standard', record.standard.name if record.standard else ''],
            ['Status', record.status],
            ['Priority', record.priority or ''],
            ['Batch Number', record.batch_number or ''],
            ['Product ID', record.product_id or ''],
            ['Location', record.location or ''],
            ['Department', record.department or ''],
            ['Created By', record.creator.full_name if record.creator else ''],
            ['Assigned To', record.assignee.full_name if record.assignee else ''],
            ['Scheduled Date', record.scheduled_date.strftime('%Y-%m-%d %H:%M') if record.scheduled_date else ''],
            ['Completed Date', record.completed_at.strftime('%Y-%m-%d %H:%M') if record.completed_at else ''],
            ['Overall Compliance', 'Pass' if record.overall_compliance else 'Fail' if record.overall_compliance is not None else ''],
            ['Compliance Score', f"{record.compliance_score}%" if record.compliance_score else ''],
            ['Failed Items', record.failed_items_count or 0],
            ['Notes', record.notes or '']
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row_idx, column=1).value = label
            ws_summary.cell(row=row_idx, column=1).font = Font(bold=True)
            ws_summary.cell(row=row_idx, column=2).value = value
        
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 50
        
        # Sheet 2: Record Items (Detailed Results)
        ws_items = wb.create_sheet("Items")
        
        item_headers = [
            'Criteria Code', 'Criteria Title', 'Value', 'Numeric Value',
            'Compliance', 'Deviation', 'Unit', 'Remarks',
            'Measured At', 'Measured By', 'Equipment'
        ]
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col, header in enumerate(item_headers, 1):
            cell = ws_items.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        for row_idx, item in enumerate(record.items, 2):
            item_data = [
                item.criteria.code if item.criteria else '',
                item.criteria.title if item.criteria else '',
                item.value or '',
                float(item.numeric_value) if item.numeric_value else '',
                'Pass' if item.compliance else 'Fail' if item.compliance is not None else '',
                float(item.deviation) if item.deviation else '',
                item.criteria.unit if item.criteria else '',
                item.remarks or '',
                item.measured_at.strftime('%Y-%m-%d %H:%M') if item.measured_at else '',
                item.measured_by.full_name if item.measured_by else '',
                item.equipment_used or ''
            ]
            
            for col, value in enumerate(item_data, 1):
                cell = ws_items.cell(row=row_idx, column=col)
                cell.value = value
                
                # Compliance color coding
                if col == 5:
                    if value == 'Pass':
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif value == 'Fail':
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        for col in range(1, len(item_headers) + 1):
            ws_items.column_dimensions[get_column_letter(col)].width = 15
        
        wb.save(filepath)
        return filepath
    
    def export_record_data(self, record: Record, filepath: str) -> str:
        """
        Export record data (criteria values only) with statistics to Excel
        No IDs, no template data - just critera names, values, and statistics
        
        Args:
            record: Record object
            filepath: Output file path
            
        Returns:
            Path to created file
        """
        import numpy as np
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Record Data"
        
        # Style definitions
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Helper function to format numbers
        def format_number(num):
            if num is None:
                return ''
            if num == int(num):
                return str(int(num))
            else:
                return f'{num:.10g}'
        
        # Title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = f"Record Data: {record.record_number}"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Record info
        ws['A2'] = "Title:"
        ws['B2'] = record.title or ''
        ws['A3'] = "Template:"
        ws['B3'] = record.template.name if record.template else ''
        ws['A4'] = "Date:"
        ws['B4'] = record.completed_at.strftime('%Y-%m-%d %H:%M') if record.completed_at else ''
        
        # Group record items by criteria
        from collections import defaultdict
        criteria_groups = defaultdict(list)
        for item in record.items:
            if item.criteria:
                criteria_groups[item.criteria].append(item)
        
        # Sort criteria based on template order if possible
        if record.template and record.template.fields:
            field_order = {f.criteria_id: f.sort_order for f in record.template.fields}
            sorted_criteria = sorted(criteria_groups.keys(), key=lambda c: field_order.get(c.id, 999))
        else:
            sorted_criteria = sorted(criteria_groups.keys(), key=lambda c: c.title or "")

        # Labels for statistics in Column A
        ws['A6'] = "STATISTICS"
        ws['A6'].font = Font(bold=True, size=12)
        ws['A7'] = "Unit"
        ws['A8'] = "Min Limit"
        ws['A9'] = "Max Limit"
        ws['A10'] = "Sample Count"
        ws['A11'] = "Average"
        ws['A12'] = "Minimum"
        ws['A13'] = "Maximum"
        ws['A14'] = "Range"
        ws['A15'] = "Std Deviation"
        
        for row in range(7, 16):
            cell = ws[f'A{row}']
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            cell.border = border

        # Data starts from Row 6 across columns B, C, D...
        col_idx = 2
        for crit in sorted_criteria:
            col_letter = get_column_letter(col_idx)
            
            # Criteria Header
            header_cell = ws[f'{col_letter}6']
            header_cell.value = crit.title
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            header_cell.border = border
            
            items = criteria_groups[crit]
            numeric_vals = [float(item.numeric_value) for item in items if item.numeric_value is not None]
            
            # Unit & Limits
            ws[f'{col_letter}7'] = crit.unit or '-'
            ws[f'{col_letter}8'] = format_number(float(crit.limit_min)) if crit.limit_min is not None else '-'
            ws[f'{col_letter}9'] = format_number(float(crit.limit_max)) if crit.limit_max is not None else '-'
            
            # Calculate and set statistics
            if numeric_vals:
                v_arr = np.array(numeric_vals)
                ws[f'{col_letter}10'] = len(numeric_vals)
                ws[f'{col_letter}11'] = format_number(np.mean(v_arr))
                ws[f'{col_letter}12'] = format_number(np.min(v_arr))
                ws[f'{col_letter}13'] = format_number(np.max(v_arr))
                ws[f'{col_letter}14'] = format_number(np.max(v_arr) - np.min(v_arr))
                ws[f'{col_letter}15'] = format_number(np.std(v_arr, ddof=1) if len(numeric_vals) > 1 else 0)
            else:
                for row in range(10, 16):
                    ws[f'{col_letter}{row}'] = 'N/A'
            
            # Apply borders to stats cells
            for row in range(7, 16):
                ws[f'{col_letter}{row}'].border = border
                ws[f'{col_letter}{row}'].alignment = Alignment(horizontal='center')

            # Readings Section
            reading_header = ws[f'{col_letter}17']
            reading_header.value = "READINGS"
            reading_header.font = Font(bold=True, color='366092')
            reading_header.alignment = Alignment(horizontal='center')
            
            # Individual reading values
            curr_row = 18
            for item in items:
                cell = ws.cell(row=curr_row, column=col_idx)
                if item.numeric_value is not None:
                    cell.value = float(item.numeric_value)
                else:
                    cell.value = item.value or ''
                
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                
                # Color code compliance
                if item.compliance is False:
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                elif item.compliance is True:
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                
                curr_row += 1
            
            # Set column width
            ws.column_dimensions[col_letter].width = max(15, len(crit.title) / 2 + 5)
            col_idx += 1
            
        ws.column_dimensions['A'].width = 20
        
        wb.save(filepath)
        return filepath
    
    def export_template_to_excel(self, template: TestTemplate, filepath: str) -> str:
        """
        Export template structure to Excel (can be used as a form)
        
        Args:
            template: TestTemplate object
            filepath: Output file path
            
        Returns:
            Path to created file
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = template.name[:31]  # Excel sheet name limit
        
        # Template header
        ws['A1'] = 'Template'
        ws['B1'] = template.name
        ws['A2'] = 'Code'
        ws['B2'] = template.code
        ws['A3'] = 'Description'
        ws['B3'] = template.description or ''
        ws['A4'] = 'Category'
        ws['B4'] = template.category or ''
        
        ws['A1'].font = Font(bold=True)
        ws['A2'].font = Font(bold=True)
        ws['A3'].font = Font(bold=True)
        ws['A4'].font = Font(bold=True)
        
        # Criteria headers
        start_row = 6
        headers = [
            'Section', 'Code', 'Title', 'Description', 'Type',
            'Required', 'Min', 'Max', 'Unit', 'Value', 'Compliance', 'Remarks'
        ]
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Template fields
        for row_idx, field in enumerate(sorted(template.fields, key=lambda x: x.sort_order or 0), start_row + 1):
            criteria = field.criteria
            data = [
                field.section_key or '',
                criteria.code,
                criteria.title,
                criteria.description or '',
                criteria.data_type,
                'Yes' if field.is_required else 'No',
                float(criteria.limit_min) if criteria.limit_min else '',
                float(criteria.limit_max) if criteria.limit_max else '',
                criteria.unit or '',
                '',  # Value - to be filled
                '',  # Compliance - to be filled
                ''   # Remarks - to be filled
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row_idx, column=col).value = value
        
        # Auto-adjust columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        wb.save(filepath)
        return filepath
    
    # ========================================================================
    # IMPORT METHODS
    # ========================================================================
    
    def import_standards_from_excel(self, filepath: str, created_by_id: int) -> List[Standard]:
        """
        Import standards from Excel file
        
        Expected columns: Code, Name, Version, Industry, Description, Effective Date
        
        Args:
            filepath: Excel file path
            created_by_id: User ID creating the standards
            
        Returns:
            List of created Standard objects
        """
        df = pd.read_excel(filepath)
        required_columns = ['Code', 'Name', 'Version']
        
        # Validate columns
        for col in required_columns:
            if col not in df.columns:
                raise ValueError(f"Missing required column: {col}")
        
        standards = []
        for _, row in df.iterrows():
            # Check if standard already exists
            existing = self.session.query(Standard).filter_by(
                code=row['Code'],
                version=row['Version']
            ).first()
            
            if existing:
                print(f"Standard {row['Code']} v{row['Version']} already exists, skipping...")
                continue
            
            standard = Standard(
                code=row['Code'],
                name=row['Name'],
                version=row['Version'],
                industry=row.get('Industry', ''),
                description=row.get('Description', ''),
                effective_date=pd.to_datetime(row['Effective Date']).date() if 'Effective Date' in row and pd.notna(row['Effective Date']) else None,
                is_active=True,
                created_by_id=created_by_id
            )
            
            self.session.add(standard)
            standards.append(standard)
        
        self.session.commit()
        return standards
    
    def import_criteria_from_excel(self, filepath: str, standard_id: int) -> List[StandardCriteria]:
        """
        Import criteria from Excel file for a specific standard
        
        Expected columns: Code, Title, Description, Type, Min, Max, Unit, Severity
        
        Args:
            filepath: Excel file path
            standard_id: Standard ID to associate criteria with
            
        Returns:
            List of created StandardCriteria objects
        """
        df = pd.read_excel(filepath)
        required_columns = ['Code', 'Title', 'Type']
        
        for col in required_columns:
            if col not in df.columns:
                raise ValueError(f"Missing required column: {col}")
        
        criteria_list = []
        for _, row in df.iterrows():
            criteria = StandardCriteria(
                standard_id=standard_id,
                code=row['Code'],
                title=row['Title'],
                description=row.get('Description', ''),
                data_type=row['Type'],
                requirement_type=row.get('Requirement Type', 'mandatory'),
                limit_min=row.get('Min') if pd.notna(row.get('Min')) else None,
                limit_max=row.get('Max') if pd.notna(row.get('Max')) else None,
                unit=row.get('Unit', ''),
                severity=row.get('Severity', 'minor'),
                is_active=True
            )
            
            self.session.add(criteria)
            criteria_list.append(criteria)
        
        self.session.commit()
        return criteria_list
    
    def import_record_from_filled_template(self, filepath: str, template_id: int, 
                                          created_by_id: int, **record_kwargs) -> Record:
        """
        Import a record from a filled template Excel file
        
        Args:
            filepath: Excel file path (filled template)
            template_id: Template ID
            created_by_id: User ID creating the record
            **record_kwargs: Additional record fields (title, batch_number, etc.)
            
        Returns:
            Created Record object
        """
        df = pd.read_excel(filepath, skiprows=5)  # Skip template header
        
        template = self.session.query(TestTemplate).get(template_id)
        if not template:
            raise ValueError(f"Template {template_id} not found")
        
        # Generate record number
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        record_number = f"REC-{timestamp}"
        
        # Create record
        record = Record(
            record_number=record_number,
            template_id=template_id,
            standard_id=template.standard_id,
            created_by_id=created_by_id,
            status='draft',
            category=template.category,
            **record_kwargs
        )
        
        self.session.add(record)
        self.session.flush()  # Get record ID
        
        # Import record items
        passed = 0
        failed = 0
        
        for _, row in df.iterrows():
            if pd.isna(row.get('Code')):
                continue
            
            # Find criteria by code
            criteria = self.session.query(StandardCriteria).filter_by(
                standard_id=template.standard_id,
                code=row['Code']
            ).first()
            
            if not criteria:
                continue
            
            # Get value
            value = str(row.get('Value', '')) if pd.notna(row.get('Value')) else None
            
            # Determine compliance
            compliance = None
            if 'Compliance' in row and pd.notna(row['Compliance']):
                compliance = row['Compliance'].lower() in ['pass', 'yes', 'true', '1']
            elif value and criteria.data_type == 'numeric':
                try:
                    numeric_val = float(value)
                    if criteria.limit_min and criteria.limit_max:
                        compliance = criteria.limit_min <= numeric_val <= criteria.limit_max
                except:
                    pass
            
            if compliance is True:
                passed += 1
            elif compliance is False:
                failed += 1
            
            # Create record item
            item = RecordItem(
                record_id=record.id,
                criteria_id=criteria.id,
                value=value,
                numeric_value=float(value) if value and criteria.data_type == 'numeric' else None,
                compliance=compliance,
                remarks=str(row.get('Remarks', '')) if pd.notna(row.get('Remarks')) else None
            )
            
            self.session.add(item)
        
        # Update record summary
        total = passed + failed
        if total > 0:
            record.compliance_score = (passed / total) * 100
            record.overall_compliance = failed == 0
            record.failed_items_count = failed
        
        self.session.commit()
        return record


# ========================================================================
# CONVENIENCE FUNCTIONS
# ========================================================================

def export_records(session, records, output_path):
    """Quick export of records"""
    handler = ExcelHandler(session)
    return handler.export_records_to_excel(records, output_path)


def export_record_detail(session, record, output_path):
    """Quick export of single record details"""
    handler = ExcelHandler(session)
    return handler.export_record_details_to_excel(record, output_path)


def export_template_to_excel(template, session, filepath):
    """
    Export template structure to Excel
    
    Args:
        template: TestTemplate object
        session: SQLAlchemy session
        filepath: Output file path
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    
    # Template Info
    ws['A1'] = 'Template Information'
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A3'] = 'Code:'
    ws['B3'] = template.code
    ws['A4'] = 'Name:'
    ws['B4'] = template.name
    ws['A5'] = 'Version:'
    ws['B5'] = template.version
    ws['A6'] = 'Category:'
    ws['B6'] = template.category
    ws['A7'] = 'Description:'
    ws['B7'] = template.description or ''
    ws['A8'] = 'Requires Approval:'
    ws['B8'] = 'Yes' if template.requires_approval else 'No'
    
    # Template Fields
    ws['A10'] = 'Template Fields'
    ws['A10'].font = Font(bold=True, size=12)
    
    # Headers for fields
    headers = ['Order', 'Criteria Code', 'Title', 'Data Type', 'Requirement', 'Required', 'Visible', 'Min', 'Max', 'Unit']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=11, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Get template fields
    from models import TemplateField
    fields = session.query(TemplateField).filter_by(
        template_id=template.id
    ).order_by(TemplateField.sort_order).all()
    
    # Write field data
    for row_num, field in enumerate(fields, 12):
        if field.criteria:
            ws.cell(row=row_num, column=1, value=field.sort_order)
            ws.cell(row=row_num, column=2, value=field.criteria.code)
            ws.cell(row=row_num, column=3, value=field.criteria.title)
            ws.cell(row=row_num, column=4, value=field.criteria.data_type)
            ws.cell(row=row_num, column=5, value=field.criteria.requirement_type)
            ws.cell(row=row_num, column=6, value='Yes' if field.is_required else 'No')
            ws.cell(row=row_num, column=7, value='Yes' if field.is_visible else 'No')
            ws.cell(row=row_num, column=8, value=str(field.criteria.limit_min) if field.criteria.limit_min else '')
            ws.cell(row=row_num, column=9, value=str(field.criteria.limit_max) if field.criteria.limit_max else '')
            ws.cell(row=row_num, column=10, value=field.criteria.unit or '')
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filepath)
    return filepath
